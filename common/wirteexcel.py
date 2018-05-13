# encoding: utf-8

from openpyxl import load_workbook
import openpyxl


def copy_excel(excelpath1, excelpath2):
    # 创建新的空的excel
    wb2 = openpyxl.Workbook()
    wb2.save(excelpath2)
    # 读取数据
    wb1 = load_workbook(excelpath1)
    wb2 = load_workbook(excelpath2)
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    sheet1 = wb1[sheets1[0]]
    sheet2 = wb2[sheets2[0]]
    max_row = sheet1.max_row
    max_column = sheet1.max_column

    for m in list(range(1, max_row+1)):
        for n in list(range(97, 97 + max_column)):
            n = chr(n)
            i = '%s%d' % (n, m)
            cell = sheet1[i].value
            sheet2[i].value = cell

    wb2.save(excelpath2)
    wb1.close()
    wb2.close()


class Write_excel(object):
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active

    def write(self, row_n, col_n, value):
        self.ws.cell(row_n, col_n).value = value
        self.wb.save(self.filename)


if __name__ == '__main__':
    copy_excel("D:\\test\\testcase.xlsx", "D:\\testreport.xlsx")
    wt = Write_excel("D:\\testreport.xlsx")
    wt.write(4, 5, "HELLEOP")
    wt.write(4, 6, "HELLEOP")
